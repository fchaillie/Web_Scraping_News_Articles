import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter
from datetime import datetime
import argparse
import os
import time

def main():

    parser = argparse.ArgumentParser(description='Download news from OKX exchange.')
    parser.add_argument('start_date', type=str, help='Start date in YYYY-MM-DD format')
    parser.add_argument('end_date', type=str, help='End date in YYYY-MM-DD format')
    parser.add_argument('folder', type=str, help='Folder to save the news data')
        
    args = parser.parse_args()

    # The dates arguments need to have the right format    
    try:
        datetime.strptime(args.start_date, '%Y-%m-%d')
        datetime.strptime(args.end_date, '%Y-%m-%d')
    except ValueError:
        raise ValueError("Dates must be in YYYY-MM-DD format")        

    # 4 lists to store the coming data
    titles = []
    date1s = []
    date2s = []
    links = []

    # URL of the website page for announcements
    base_url = "https://www.okx.com/help/section/announcements-latest-announcements"
    page_number = 1  # Start on page 1

    # Set this variable to stop the search once the start date is passed
    start_date_not_passed = True

    while start_date_not_passed:

        # Give the loop 1 second before each iteration
        time.sleep(1)

        # URL to target
        if page_number == 1:
            url = base_url
        else:
            url = f"{base_url}/page/{page_number}"

        response = requests.get(url)
        # Only if the page is reached do we keep going
        if response.status_code != 200:
            break
        # Get all the articles from the HTML page in a list
        soup = BeautifulSoup(response.content, 'html.parser')
        articles = list(soup.find('div', class_= "index_list__lAqHy index_list__yJClY"))

        # For each article, we get the title, the publication date, the update date and the link
        for article in articles:

            title_div = article.find('div', class_='index_articleTitle__ys7G7')
            if title_div:
                # Title and link
                title = article.find('div', class_='index_articleTitle__ys7G7').text.strip()
                link = article['href']
                link = f"https://www.okx.com{link}"

                date_div = article.find('div', class_='index_dividerRow__FHkzs index_detailsRow__8Gmjm')
                if date_div:

                    # Find all the dates for each article, there are up to 2 dates
                    dates = date_div.find_all('span', attrs={'data-testid': 'DateDisplay'})

                    # Publication date
                    date_pusblished = dates[0].text.strip()
                    date_pusblished = date_pusblished.split("Published on ")
                    date_pusblished = date_pusblished[1]

                    # Format publication date to compare it to start_date and end_date
                    date_obj = datetime.strptime(date_pusblished, '%b %d, %Y')
                    formatted_date = date_obj.strftime('%Y-%m-%d')
                    # If the publication date is more recent than the end_date than we go to the next article
                    if args.end_date < formatted_date:
                        continue
                    # If the publication date is older than the start_date than we stop searching articles
                    if args.start_date > formatted_date:
                        start_date_not_passed = False
                        break
                    # If there is a second date than it is the update date
                    if len(dates) > 1:
                        date_update = dates[1].text.strip()  # Second date
                        date_update = date_update.split("Updated on ")
                        date_update = date_update[1]
                        # The update date will be shown if different from the publication date
                        if date_update == date_pusblished:
                            date_update = ' '
                            date2s.append(date_update)
                        else:
                            date2s.append(date_update)
                    titles.append(title)
                    date1s.append(date_pusblished)   
                    links.append(link)

        page_number += 1
                    
    # Convert the 4 lists into a DataFrame
    df = pd.DataFrame({
        'Title': titles,
        'Publication Date': date1s,
        'Update Date': date2s,
        'Link': links
        })
    # If the DataFrame is empty then no file is created, only a message
    if df.shape[0] == 0:
        print("There are no articles available for this time period")

    else:
        # If the folder for the excel file is not created yet then it gets created    
        if not os.path.isdir(args.folder):
            os.mkdir(args.folder)
        
        # Name of the excel file created included start_date and end_date
        excel_path = f'{args.folder}/Articles_from_{args.start_date}_to_{args.end_date}.xlsx'

        # Insert the 4 lists to the Excel file in the first 4 columns
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df['Title'].to_excel(writer, startcol=0, index=False, sheet_name='Sheet1')  # Column 1
            df['Publication Date'].to_excel(writer, startcol=1, index=False, sheet_name='Sheet1') # Column 2
            df['Update Date'].to_excel(writer, startcol=2, index=False, sheet_name='Sheet1') # Column 3
            df['Link'].to_excel(writer, startcol=3, index=False, sheet_name='Sheet1')  # Column 4


        # Load the workbook and select the active worksheet
        wb = load_workbook(excel_path)
        ws = wb.active

        # Iterate through each cell in the link column to create a clickable link , skipping the header row of course
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, 4)
            if cell.value:
                # Create the clickable hyperlink
                cell.hyperlink = cell.value
                cell.value = "Click to view"
                cell.style = "Hyperlink"
                
        # Adjust width of each column
        for col in ws.columns:
            max_length = 0
            # Get the column index number
            column = col[0].column  

            for cell in col:
                # Check if the cell text is longer than the current max length
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
               
            # Add some extra space
            adjusted_width = (max_length + 2)  
            # Set the final width of each column
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        wb.save(excel_path)

        print("Excel file has been created with article titles, dates and links.")

# Activate the main function only if this script is executed directly
if __name__ == "__main__":
    main()
